"""Excel 結構完整性驗證。

用法：
    python verify_structure.py <xlsx_path>

回傳 JSON 報告，包含所有結構元素的盤點。
用於驗證產出的 xlsx 是否包含所有預期的結構元素。
"""

import json
import sys
from openpyxl import load_workbook


def verify_structure(filepath):
    """讀取 xlsx 並回傳完整結構盤點報告。"""
    wb = load_workbook(filepath, data_only=False)
    report = {
        'sheets': [],
        'defined_names': [],
        'metadata': {
            'title': wb.properties.title,
            'creator': wb.properties.creator,
        },
    }

    # 命名範圍
    if wb.defined_names:
        for dn in wb.defined_names.definedName:
            report['defined_names'].append({
                'name': dn.name,
                'value': dn.attr_text,
                'scope': dn.localSheetId,
            })

    # 逐 Sheet 盤點
    for ws in wb.worksheets:
        sheet_info = {
            'name': ws.title,
            'state': ws.sheet_state,
            'dimensions': ws.calculate_dimension(),
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'freeze_panes': ws.freeze_panes,
            'print_area': ws.print_area,
            'print_title_rows': ws.print_title_rows,
            'print_title_cols': ws.print_title_cols,
            'merged_cells': [str(m) for m in ws.merged_cells.ranges],
            'auto_filter': ws.auto_filter.ref if ws.auto_filter.ref else None,
            'data_validations': [],
            'conditional_formats': [],
            'tables': [],
            'hidden_rows': [],
            'hidden_cols': [],
            'protection': {
                'enabled': ws.protection.sheet,
            },
            'page_setup': {
                'orientation': ws.page_setup.orientation,
                'paper_size': ws.page_setup.paperSize,
            },
        }

        # 資料驗證
        for dv in ws.data_validations.dataValidation:
            sheet_info['data_validations'].append({
                'range': str(dv.sqref),
                'type': dv.type,
                'formula1': dv.formula1,
                'operator': dv.operator,
            })

        # 條件格式
        for range_str, rules in ws.conditional_formatting:
            for rule in rules:
                sheet_info['conditional_formats'].append({
                    'range': str(range_str),
                    'type': rule.type,
                    'priority': rule.priority,
                })

        # 表格
        for table in ws.tables.values():
            sheet_info['tables'].append({
                'name': table.name,
                'ref': table.ref,
                'columns': [c.name for c in table.tableColumns],
            })

        # 隱藏行列
        for row_idx, rd in ws.row_dimensions.items():
            if rd.hidden:
                sheet_info['hidden_rows'].append(row_idx)
        for col_letter, cd in ws.column_dimensions.items():
            if cd.hidden:
                sheet_info['hidden_cols'].append(col_letter)

        report['sheets'].append(sheet_info)

    wb.close()
    return report


def compare_with_spec(report, spec):
    """比對結構報告與預期規格，回傳差異清單。

    spec 格式範例：
    {
        "sheets": ["Sheet1", "Sheet2"],
        "freeze_panes": {"Sheet1": "A3"},
        "merged_cells": {"Sheet1": ["A1:D1", "A2:D2"]},
        "data_validations": [{"sheet": "Sheet1", "range": "B5:B20", "type": "list"}],
    }
    """
    errors = []

    # Sheet 存在性
    actual_sheets = [s['name'] for s in report['sheets']]
    for name in spec.get('sheets', []):
        if name not in actual_sheets:
            errors.append(f"缺少 Sheet: {name}")

    # 凍結窗格
    for sheet_name, expected_freeze in spec.get('freeze_panes', {}).items():
        for s in report['sheets']:
            if s['name'] == sheet_name:
                if s['freeze_panes'] != expected_freeze:
                    errors.append(
                        f"{sheet_name}: 凍結窗格應為 {expected_freeze}，實為 {s['freeze_panes']}"
                    )

    # 合併儲存格
    for sheet_name, expected_merges in spec.get('merged_cells', {}).items():
        for s in report['sheets']:
            if s['name'] == sheet_name:
                for m in expected_merges:
                    if m not in s['merged_cells']:
                        errors.append(f"{sheet_name}: 缺少合併儲存格 {m}")

    # 資料驗證
    for edv in spec.get('data_validations', []):
        found = False
        for s in report['sheets']:
            if s['name'] == edv['sheet']:
                for dv in s['data_validations']:
                    if dv['range'] == edv['range'] and dv['type'] == edv['type']:
                        found = True
                        break
        if not found:
            errors.append(f"{edv['sheet']} {edv['range']}: 缺少資料驗證 ({edv['type']})")

    return errors


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('用法: python verify_structure.py <xlsx_path> [spec.json]')
        sys.exit(1)

    report = verify_structure(sys.argv[1])

    if len(sys.argv) >= 3:
        with open(sys.argv[2], 'r', encoding='utf-8') as f:
            spec = json.load(f)
        errors = compare_with_spec(report, spec)
        report['comparison_errors'] = errors

    print(json.dumps(report, ensure_ascii=False, indent=2))
