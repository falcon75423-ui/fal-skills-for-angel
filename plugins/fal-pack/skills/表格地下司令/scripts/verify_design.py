#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""verify_design.py — 設計品味地雷自驗腳本

F-018 借鑑 impeccable v3.0.5 E19「Build-time dogfood validator」精神。
對應 landmines.md「## 3. 設計品味地雷」8 條 ❌→✅ 對照。

接到驗證管線 Layer 1.5 位置（結構驗證後、視覺品檢前）。

維度（6 個 check）：
  1. 字型計數 ≤ 3 套（含 Latin + CJK + 等寬數字全部）
  2. 隔行底色灰度差（≤4% INFO / 4-8% WARNING / >8% BLOCKER）
  3. 列高 6pt 整數倍（15pt 預設為合法例外 → WARNING；
     合併儲存格 + > 50pt 視為動態列高合法例外。
     v1 限制：只檢查合併，未檢查 wrap_text；標 v2 待辦——
     合併但未 wrap 的 60pt 列會被誤判為合法例外）
  4. 數字欄一致對齊（同欄混用左右對齊 → BLOCKER）
  5. 紅綠對比偵測（單純紅綠無形狀差異 → BLOCKER）
  6. 純黑文字 + 純白背景（同時出現 → WARNING）

輸出分級：
  🔴 BLOCKER：違反根本原則，必修才交付
  🟡 WARNING：有風險或合法例外，需要人工判斷
  🟢 INFO：通過 / 統計訊息

退出碼：
  0 = 全綠（無 BLOCKER 無 WARNING）
  1 = 有 WARNING 無 BLOCKER
  2 = 有 BLOCKER

用法：
    python verify_design.py path/to/file.xlsx
    python verify_design.py path/to/file.xlsx --json   # 機器讀格式
    python verify_design.py path/to/file.xlsx --strict # WARNING 也視為 fail

來源：F-018 火神鍛造工頭 + 表哥（表格地下司令）跨客戶協作鍛造（2026-04-30）。
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    sys.stderr.write("ERROR: openpyxl 未安裝。執行: pip install openpyxl\n")
    sys.exit(1)


# ============================================================
# 常數
# ============================================================

MAX_FONTS = 3
ZEBRA_GRAY_DIFF_INFO = 4.0      # %  ≤ 4% = 通過
ZEBRA_GRAY_DIFF_WARN = 8.0      # %  4-8% = WARNING
# > 8% = BLOCKER

LINE_HEIGHT_BASE = 6            # pt 基礎單位（跟 F-017 對齊）
LINE_HEIGHT_LEGAL_EXCEPTION = {15.0}  # Excel 預設 15pt 合法例外
DYNAMIC_HEIGHT_THRESHOLD = 50.0  # > 50pt + merged cell = dynamic row exemption (F-018 B-3)

# 紅綠色相邊界（HSV）
RED_HUE_RANGES = ((0, 15), (345, 360))   # 紅色 hue 範圍
GREEN_HUE_RANGES = ((90, 150),)          # 綠色 hue 範圍

# 形狀差異標記（紅綠對比若同時有這些字符 = 視為合法）
SHAPE_MARKERS = {"▲", "▼", "▴", "▾", "↑", "↓", "+", "-"}

PURE_BLACK = "FF000000"
PURE_WHITE = "FFFFFFFF"


# ============================================================
# Finding 結構
# ============================================================

def make_finding(level: str, code: str, msg: str, ref: str = "") -> dict:
    """level ∈ {BLOCKER, WARNING, INFO}"""
    return {"level": level, "code": code, "msg": msg, "ref": ref}


# ============================================================
# 工具函式
# ============================================================

def rgb_to_gray_pct(rgb_hex: str) -> float | None:
    """RGB hex → 灰度（0-100%）。'FF' 開頭的 alpha 跳過。"""
    if not rgb_hex or not isinstance(rgb_hex, str):
        return None
    s = rgb_hex.upper().lstrip("#")
    if len(s) == 8:
        s = s[2:]  # 去 alpha
    if len(s) != 6:
        return None
    try:
        r = int(s[0:2], 16)
        g = int(s[2:4], 16)
        b = int(s[4:6], 16)
    except ValueError:
        return None
    # ITU-R BT.601 luminance
    gray = 0.299 * r + 0.587 * g + 0.114 * b
    return gray / 255.0 * 100.0


def rgb_to_hsv(rgb_hex: str) -> tuple[float, float, float] | None:
    """RGB hex → HSV (h:0-360, s:0-1, v:0-1)。"""
    if not rgb_hex or not isinstance(rgb_hex, str):
        return None
    s = rgb_hex.upper().lstrip("#")
    if len(s) == 8:
        s = s[2:]
    if len(s) != 6:
        return None
    try:
        r = int(s[0:2], 16) / 255.0
        g = int(s[2:4], 16) / 255.0
        b = int(s[4:6], 16) / 255.0
    except ValueError:
        return None
    mx = max(r, g, b)
    mn = min(r, g, b)
    diff = mx - mn
    if diff == 0:
        h = 0.0
    elif mx == r:
        h = (60 * ((g - b) / diff) + 360) % 360
    elif mx == g:
        h = 60 * ((b - r) / diff) + 120
    else:
        h = 60 * ((r - g) / diff) + 240
    sat = 0.0 if mx == 0 else diff / mx
    return h, sat, mx


def hue_in_ranges(hue: float, ranges) -> bool:
    return any(lo <= hue <= hi for (lo, hi) in ranges)


def get_cell_color_rgb(cell, attr: str) -> str | None:
    """從 cell.font.color / cell.fill.start_color 取 rgb 字串。
    attr 是 'font' 或 'fill'。Theme Color 回 None（已在相容性地雷禁止）。
    """
    try:
        if attr == "font":
            color = cell.font.color
        elif attr == "fill":
            color = cell.fill.start_color if cell.fill else None
        else:
            return None
        if color is None:
            return None
        # color.type 'rgb' / 'theme' / 'indexed'
        if getattr(color, "type", None) == "rgb":
            return color.rgb
        return None
    except (AttributeError, TypeError):
        return None


def get_cell_text(cell) -> str:
    v = cell.value
    return "" if v is None else str(v)


# ============================================================
# Check 1: 字型計數 ≤ 3
# ============================================================

def check_font_count(wb: Workbook) -> list[dict]:
    """掃所有 cell 的 font.name，超過 3 套 = BLOCKER。"""
    fonts: set[str] = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.font and cell.font.name:
                    fonts.add(cell.font.name)

    findings = []
    if not fonts:
        findings.append(make_finding("INFO", "FONT_COUNT", "未偵測到任何字型 (空檔?)"))
    else:
        font_list = sorted(fonts)
        if len(fonts) <= MAX_FONTS:
            findings.append(make_finding(
                "INFO", "FONT_COUNT",
                f"字型 {len(fonts)} 套 (≤ {MAX_FONTS}): {', '.join(font_list)}",
                ref="landmines.md § 3 設計品味地雷 #3"
            ))
        else:
            findings.append(make_finding(
                "BLOCKER", "FONT_COUNT",
                f"字型 {len(fonts)} 套 (> {MAX_FONTS}): {', '.join(font_list)}",
                ref="landmines.md § 3 設計品味地雷 #3"
            ))
    return findings


# ============================================================
# Check 2: 隔行底色灰度差
# ============================================================

def check_zebra_contrast(ws: Worksheet) -> list[dict]:
    """偵測連續資料列若用 zebra striping，灰度差是否合理。

    工法：抽樣前 20 列（跳過第 1 列假設為標題），算奇偶列底色灰度差均值。
    """
    if ws.max_row < 4:
        return []  # 少於 4 列無法判斷

    odd_grays: list[float] = []
    even_grays: list[float] = []
    sample_rows = min(20, ws.max_row - 1)

    for r in range(2, 2 + sample_rows):
        # 抽樣每列 col 2 的 fill 顏色作代表
        cell = ws.cell(row=r, column=2)
        rgb = get_cell_color_rgb(cell, "fill")
        gray = rgb_to_gray_pct(rgb) if rgb else None
        if gray is None:
            continue
        if (r - 1) % 2 == 1:
            odd_grays.append(gray)
        else:
            even_grays.append(gray)

    if not odd_grays or not even_grays:
        return [make_finding(
            "INFO", "ZEBRA_CONTRAST",
            f"工作表 '{ws.title}' 未偵測到隔行底色 (跳過檢查)"
        )]

    odd_avg = sum(odd_grays) / len(odd_grays)
    even_avg = sum(even_grays) / len(even_grays)
    diff = abs(odd_avg - even_avg)

    findings = []
    if diff <= ZEBRA_GRAY_DIFF_INFO:
        findings.append(make_finding(
            "INFO", "ZEBRA_CONTRAST",
            f"工作表 '{ws.title}' 隔行底色差 {diff:.1f}% (≤ {ZEBRA_GRAY_DIFF_INFO}% 通過)",
            ref="landmines.md § 3 設計品味地雷 #6"
        ))
    elif diff <= ZEBRA_GRAY_DIFF_WARN:
        findings.append(make_finding(
            "WARNING", "ZEBRA_CONTRAST",
            f"工作表 '{ws.title}' 隔行底色差 {diff:.1f}% ({ZEBRA_GRAY_DIFF_INFO}-{ZEBRA_GRAY_DIFF_WARN}% 警告)",
            ref="landmines.md § 3 設計品味地雷 #6"
        ))
    else:
        findings.append(make_finding(
            "BLOCKER", "ZEBRA_CONTRAST",
            f"工作表 '{ws.title}' 隔行底色差 {diff:.1f}% (> {ZEBRA_GRAY_DIFF_WARN}% 搶 hierarchy)",
            ref="landmines.md § 3 設計品味地雷 #6"
        ))
    return findings


# ============================================================
# Check 3: 列高 6pt 整數倍
# ============================================================

def check_row_height_rhythm(ws: Worksheet) -> list[dict]:
    """列高應為 6pt 整數倍。15pt 預設為合法例外 (WARNING)。"""
    findings = []
    bad_rows: list[tuple[int, float]] = []
    legal_exception_rows: list[int] = []
    dynamic_height_rows: list[int] = []

    # F-018 B-3: collect merged cell row numbers for dynamic height exemption
    merged_row_set: set[int] = set()
    for merge_range in ws.merged_cells.ranges:
        for r in range(merge_range.min_row, merge_range.max_row + 1):
            merged_row_set.add(r)

    for row_idx, row_dim in ws.row_dimensions.items():
        h = row_dim.height
        if h is None:
            continue  # 預設列高跳過
        if h in LINE_HEIGHT_LEGAL_EXCEPTION:
            legal_exception_rows.append(row_idx)
            continue
        # F-018 B-3: dynamic height exemption (merged cell + > 50pt = wrap_text long content)
        if h > DYNAMIC_HEIGHT_THRESHOLD and row_idx in merged_row_set:
            dynamic_height_rows.append(row_idx)
            continue
        # 容差 0.01pt 容忍浮點誤差
        if abs(h % LINE_HEIGHT_BASE) > 0.01 and abs(h % LINE_HEIGHT_BASE - LINE_HEIGHT_BASE) > 0.01:
            bad_rows.append((row_idx, h))

    if bad_rows:
        sample = ", ".join(f"R{r}={h}pt" for r, h in bad_rows[:5])
        suffix = f" (+{len(bad_rows)-5} more)" if len(bad_rows) > 5 else ""
        findings.append(make_finding(
            "BLOCKER", "ROW_HEIGHT_RHYTHM",
            f"工作表 '{ws.title}' 有 {len(bad_rows)} 列非 {LINE_HEIGHT_BASE}pt 整倍數: {sample}{suffix}",
            ref="landmines.md § 3 設計品味地雷 #4"
        ))
    if legal_exception_rows:
        findings.append(make_finding(
            "WARNING", "ROW_HEIGHT_RHYTHM",
            f"工作表 '{ws.title}' 有 {len(legal_exception_rows)} 列用 15pt (Excel 預設合法例外)",
            ref="landmines.md § 3 設計品味地雷 #4"
        ))
    if dynamic_height_rows:
        findings.append(make_finding(
            "INFO", "ROW_HEIGHT_RHYTHM",
            f"工作表 '{ws.title}' 有 {len(dynamic_height_rows)} 列為動態列高 (合併儲存格 + > {DYNAMIC_HEIGHT_THRESHOLD}pt 合法例外)",
            ref="landmines.md § 3 設計品味地雷 #4"
        ))
    if not bad_rows and not legal_exception_rows and not dynamic_height_rows:
        findings.append(make_finding(
            "INFO", "ROW_HEIGHT_RHYTHM",
            f"工作表 '{ws.title}' 列高全為 {LINE_HEIGHT_BASE}pt 整倍數",
            ref="landmines.md § 3 設計品味地雷 #4"
        ))
    return findings


# ============================================================
# Check 4: 數字欄一致對齊
# ============================================================

def check_number_alignment(ws: Worksheet) -> list[dict]:
    """同欄全是數字時，alignment.horizontal 應一致為 'right'。
    混用 'left' / 'right' / None / 'center' = BLOCKER。
    """
    findings = []
    bad_cols: list[tuple[str, set]] = []

    for col_idx in range(1, ws.max_column + 1):
        alignments: set[str] = set()
        all_numeric = True
        cell_count = 0
        for row_idx in range(2, ws.max_row + 1):  # 跳過第 1 列假設為表頭
            cell = ws.cell(row=row_idx, column=col_idx)
            v = cell.value
            if v is None:
                continue
            if not isinstance(v, (int, float)):
                all_numeric = False
                break
            cell_count += 1
            h = cell.alignment.horizontal if cell.alignment else None
            alignments.add(h or "default")

        if all_numeric and cell_count >= 3 and len(alignments) > 1:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_idx)
            bad_cols.append((col_letter, alignments))

    if bad_cols:
        sample = ", ".join(f"{col}={a}" for col, a in bad_cols[:3])
        findings.append(make_finding(
            "BLOCKER", "NUMBER_ALIGNMENT",
            f"工作表 '{ws.title}' 有 {len(bad_cols)} 個數字欄混用對齊: {sample}",
            ref="landmines.md § 3 設計品味地雷 #7"
        ))
    else:
        findings.append(make_finding(
            "INFO", "NUMBER_ALIGNMENT",
            f"工作表 '{ws.title}' 數字欄對齊一致",
            ref="landmines.md § 3 設計品味地雷 #7"
        ))
    return findings


# ============================================================
# Check 5: 紅綠對比偵測
# ============================================================

def check_red_green_contrast(ws: Worksheet) -> list[dict]:
    """偵測同欄是否同時用「紅色字」+「綠色字」且無形狀差異標記。
    色弱讀者完全看不出 = BLOCKER。
    """
    findings = []
    bad_cols: list[str] = []

    for col_idx in range(1, ws.max_column + 1):
        has_red = False
        has_green = False
        has_shape = False
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            rgb = get_cell_color_rgb(cell, "font")
            if rgb:
                hsv = rgb_to_hsv(rgb)
                if hsv and hsv[1] >= 0.4 and hsv[2] >= 0.3:  # 飽和度+明度過濾灰
                    if hue_in_ranges(hsv[0], RED_HUE_RANGES):
                        has_red = True
                    elif hue_in_ranges(hsv[0], GREEN_HUE_RANGES):
                        has_green = True
            text = get_cell_text(cell)
            if any(m in text for m in SHAPE_MARKERS):
                has_shape = True

        if has_red and has_green and not has_shape:
            from openpyxl.utils import get_column_letter
            bad_cols.append(get_column_letter(col_idx))

    if bad_cols:
        findings.append(make_finding(
            "BLOCKER", "RED_GREEN_CONTRAST",
            f"工作表 '{ws.title}' 欄 {', '.join(bad_cols)} 用紅綠對比但無形狀標記 (▲/▼/+/-)，色弱讀者看不出",
            ref="landmines.md § 3 設計品味地雷 #1"
        ))
    else:
        findings.append(make_finding(
            "INFO", "RED_GREEN_CONTRAST",
            f"工作表 '{ws.title}' 紅綠對比安全",
            ref="landmines.md § 3 設計品味地雷 #1"
        ))
    return findings


# ============================================================
# Check 6: 純黑/純白
# ============================================================

def check_pure_black_white(ws: Worksheet) -> list[dict]:
    """純黑文字 + 純白背景同時出現 = WARNING。"""
    findings = []
    has_pure_black_text = False
    has_pure_white_bg = False

    for row in ws.iter_rows():
        for cell in row:
            font_rgb = get_cell_color_rgb(cell, "font")
            fill_rgb = get_cell_color_rgb(cell, "fill")
            if font_rgb and font_rgb.upper() == PURE_BLACK:
                has_pure_black_text = True
            if fill_rgb and fill_rgb.upper() == PURE_WHITE:
                has_pure_white_bg = True
            if has_pure_black_text and has_pure_white_bg:
                break
        if has_pure_black_text and has_pure_white_bg:
            break

    if has_pure_black_text and has_pure_white_bg:
        findings.append(make_finding(
            "WARNING", "PURE_BLACK_WHITE",
            f"工作表 '{ws.title}' 同時出現 #000 純黑文字 + #FFF 純白背景 (傷視疲勞 + 紙頁缺溫度)",
            ref="landmines.md § 3 設計品味地雷 #8"
        ))
    else:
        findings.append(make_finding(
            "INFO", "PURE_BLACK_WHITE",
            f"工作表 '{ws.title}' 未同時出現純黑+純白",
            ref="landmines.md § 3 設計品味地雷 #8"
        ))
    return findings


# ============================================================
# Main
# ============================================================

def verify(xlsx_path: str | Path) -> list[dict]:
    """跑所有 check，回傳 findings list。"""
    p = Path(xlsx_path)
    if not p.exists():
        return [make_finding("BLOCKER", "FILE_NOT_FOUND", f"檔案不存在: {p}")]

    try:
        wb = load_workbook(p, data_only=False)
    except Exception as e:
        return [make_finding("BLOCKER", "LOAD_FAIL", f"無法載入 .xlsx: {e}")]

    findings: list[dict] = []

    # 全 workbook 級
    findings.extend(check_font_count(wb))

    # 每個 worksheet
    for ws in wb.worksheets:
        findings.extend(check_zebra_contrast(ws))
        findings.extend(check_row_height_rhythm(ws))
        findings.extend(check_number_alignment(ws))
        findings.extend(check_red_green_contrast(ws))
        findings.extend(check_pure_black_white(ws))

    return findings


def format_findings(findings: list[dict]) -> str:
    """人類可讀格式輸出。"""
    LEVEL_EMOJI = {"BLOCKER": "🔴", "WARNING": "🟡", "INFO": "🟢"}
    lines = []
    for f in findings:
        emoji = LEVEL_EMOJI.get(f["level"], "")
        line = f"  {emoji} [{f['level']}] {f['code']}: {f['msg']}"
        if f.get("ref"):
            line += f"\n    ref: {f['ref']}"
        lines.append(line)
    return "\n".join(lines)


def summarize(findings: list[dict]) -> tuple[int, int, int]:
    blockers = sum(1 for f in findings if f["level"] == "BLOCKER")
    warnings = sum(1 for f in findings if f["level"] == "WARNING")
    infos = sum(1 for f in findings if f["level"] == "INFO")
    return blockers, warnings, infos


def main():
    parser = argparse.ArgumentParser(
        description="表格地下司令 verify_design.py — 設計品味地雷自驗 (F-018)"
    )
    parser.add_argument("xlsx_path", help="待驗證的 .xlsx 檔案路徑")
    parser.add_argument(
        "--json", action="store_true",
        help="輸出 JSON 格式 (機器讀)"
    )
    parser.add_argument(
        "--strict", action="store_true",
        help="WARNING 也視為 fail (exit 1)"
    )
    args = parser.parse_args()

    findings = verify(args.xlsx_path)
    blockers, warnings, infos = summarize(findings)

    if args.json:
        out = {
            "path": str(args.xlsx_path),
            "summary": {"blockers": blockers, "warnings": warnings, "infos": infos},
            "findings": findings,
        }
        print(json.dumps(out, ensure_ascii=False, indent=2))
    else:
        print(f"=== 設計品味驗證: {args.xlsx_path} ===")
        print(f"  BLOCKER: {blockers} | WARNING: {warnings} | INFO: {infos}")
        print()
        print(format_findings(findings))
        print()

    if blockers:
        sys.exit(2)
    if args.strict and warnings:
        sys.exit(1)
    sys.exit(0)


if __name__ == "__main__":
    main()
