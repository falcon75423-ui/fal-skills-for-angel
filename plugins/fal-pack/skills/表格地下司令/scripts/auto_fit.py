"""CJK-aware 欄寬列高自動計算工具。

用法：
    from auto_fit import auto_fit_columns, auto_fit_rows
    auto_fit_columns(ws, min_width=8, max_width=50, padding=2)
    auto_fit_rows(ws)
"""

import unicodedata
from openpyxl.utils import get_column_letter


def get_display_width(text):
    """計算文字的顯示寬度。CJK Fullwidth/Wide 算 2，其餘算 1。"""
    if not text:
        return 0
    width = 0
    for char in str(text):
        eaw = unicodedata.east_asian_width(char)
        if eaw in ('F', 'W'):
            width += 2
        else:
            width += 1
    return width


def auto_fit_columns(ws, min_width=8, max_width=50, padding=2, gsheets_buffer=True):
    """自動計算並設定每欄的寬度。

    Args:
        ws: openpyxl Worksheet
        min_width: 最小欄寬（字元數）
        max_width: 最大欄寬（字元數）
        padding: 額外填充（字元數）
        gsheets_buffer: 是否加 Google Sheets 相容 buffer（額外 +1）
    """
    for col_idx in range(1, ws.max_column + 1):
        max_w = 0
        col_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value is None:
                    continue

                # 多行取最寬行
                lines = str(cell.value).split('\n')
                cell_width = max(get_display_width(line) for line in lines)

                # 字型大小縮放（以 11pt 為基準）
                font_size = 11
                if cell.font and cell.font.size:
                    font_size = cell.font.size
                    cell_width = int(cell_width * (font_size / 11.0))

                # 粗體加 5%
                if cell.font and cell.font.bold:
                    cell_width = int(cell_width * 1.05)

                max_w = max(max_w, cell_width)

        adjusted = max_w + padding
        if gsheets_buffer:
            adjusted += 1  # Google Sheets 欄寬偏差 buffer

        ws.column_dimensions[col_letter].width = min(max(adjusted, min_width), max_width)


def auto_fit_rows(ws, default_height=15, line_spacing=1.2):
    """自動計算並設定 wrap_text 行的列高。

    只處理 wrap_text=True 的行。其他行保持預設。
    """
    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        font_size = 11

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue

            # 取該 cell 的字型大小
            if cell.font and cell.font.size:
                font_size = max(font_size, cell.font.size)

            # 有 wrap_text 才計算多行
            if cell.alignment and cell.alignment.wrap_text:
                content = str(cell.value)
                explicit_lines = content.count('\n') + 1

                # 估算 wrap 行數
                col_letter = get_column_letter(col_idx)
                col_width = ws.column_dimensions[col_letter].width or 8
                content_width = get_display_width(content)
                wrap_lines = max(1, -(-content_width // int(col_width)))  # ceiling div

                max_lines = max(max_lines, explicit_lines, wrap_lines)

        if max_lines > 1:
            ws.row_dimensions[row_idx].height = max(
                default_height,
                max_lines * font_size * line_spacing
            )
