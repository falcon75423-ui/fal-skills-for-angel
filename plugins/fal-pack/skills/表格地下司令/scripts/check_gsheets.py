"""Google Sheets 相容性閘門。

用法：
    python check_gsheets.py <xlsx_path>

檢查層級：
    1. openpyxl 層級：Theme Color、不安全字型、Gradient Fill、不相容公式
    2. ZIP 層級：VBA、ActiveX、Sparklines、OLE、外部連線

回傳 JSON：{"compatible": true/false, "warnings": [...]}
"""

import json
import os
import re
import sys
import zipfile


SAFE_FONTS = {
    'Arial', 'Noto Sans TC', 'Noto Sans SC', 'Noto Sans JP',
    'Roboto', 'Courier New', 'Roboto Mono', None
}

BANNED_FORMULAS = ['FILTER(', 'SORT(', 'UNIQUE(', 'SEQUENCE(', 'LAMBDA(', 'LET(', 'CUBE']


def check_openpyxl_level(filepath):
    from openpyxl import load_workbook
    warnings = []
    wb = load_workbook(filepath, data_only=False)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                # Theme Color
                if cell.font.color and cell.font.color.theme is not None:
                    warnings.append(f"{ws.title}!{cell.coordinate}: 字型使用 Theme Color（會變黑色）")
                if (cell.fill.fgColor and cell.fill.fgColor.theme is not None
                        and cell.fill.patternType and cell.fill.patternType != 'none'):
                    warnings.append(f"{ws.title}!{cell.coordinate}: 填充使用 Theme Color")

                # 不安全字型
                if cell.font.name and cell.font.name not in SAFE_FONTS:
                    warnings.append(f"{ws.title}!{cell.coordinate}: 字型 '{cell.font.name}' 非 Google Sheets 安全字型")

                # Gradient Fill
                if hasattr(cell.fill, 'type') and cell.fill.type == 'gradient':
                    warnings.append(f"{ws.title}!{cell.coordinate}: 使用 Gradient Fill（只取第一色）")

                # 不相容公式
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_upper = cell.value.upper()
                    for fn in BANNED_FORMULAS:
                        if fn in formula_upper:
                            warnings.append(
                                f"{ws.title}!{cell.coordinate}: 使用不相容公式 {fn.rstrip('(')}"
                            )

    wb.close()
    return warnings


def check_zip_level(filepath):
    warnings = []
    with zipfile.ZipFile(filepath) as z:
        names = z.namelist()

        if 'xl/vbaProject.bin' in names:
            warnings.append('致命：VBA macro 會被 Google Sheets 剝離')
        if any('activeX' in n for n in names):
            warnings.append('致命：ActiveX 控件會消失')
        if 'xl/connections.xml' in names:
            warnings.append('致命：外部資料連線會消失')
        if any('oleObject' in n.lower() for n in names):
            warnings.append('致命：OLE 物件會消失')
        if any('diagrams/' in n for n in names):
            warnings.append('高：SmartArt 會被柵格化或消失')

        for name in names:
            if name.startswith('xl/worksheets/sheet'):
                content = z.read(name).decode('utf-8', errors='ignore')
                if 'sparklineGroup' in content:
                    warnings.append('致命：Sparklines 會消失')
                    break

    # 檔案大小
    size_mb = os.path.getsize(filepath) / (1024 * 1024)
    if size_mb > 100:
        warnings.append(f'致命：檔案 {size_mb:.1f}MB 超過 Google Sheets 100MB 上限')

    return warnings


def check_gsheets_compatibility(filepath):
    warnings = []
    warnings.extend(check_zip_level(filepath))
    warnings.extend(check_openpyxl_level(filepath))

    return {
        'compatible': len(warnings) == 0,
        'warnings': warnings,
        'fatal_count': sum(1 for w in warnings if w.startswith('致命')),
        'total_count': len(warnings),
    }


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('用法: python check_gsheets.py <xlsx_path>')
        sys.exit(1)

    result = check_gsheets_compatibility(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False, indent=2))
